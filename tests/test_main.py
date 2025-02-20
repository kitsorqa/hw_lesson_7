import csv

from pypdf import PdfReader
from zipfile import ZipFile
from tests.conftest import ARCHIVE_FILE_PATH
from openpyxl import load_workbook


def test_pdf():
    with ZipFile(ARCHIVE_FILE_PATH) as zip_file:
        with zip_file.open("pdf_example.pdf") as pdf:
            pdf_reader = PdfReader(pdf)
            count_of_pages = len(pdf_reader.pages)

            assert count_of_pages > 0
            assert 'Python Testing with pytest' in pdf_reader.pages[5].extract_text()


def test_csv():
    with ZipFile(ARCHIVE_FILE_PATH) as zip_file:
        with zip_file.open("csv_example.csv") as csv_file:
            csv_content = csv_file.read().decode('utf-8')
            csv_reader = list(csv.reader(csv_content.splitlines()))

            assert csv_reader[5][1] == 'Д'
            assert len(csv_reader) == 34


def test_xlsx():
    with ZipFile(ARCHIVE_FILE_PATH) as zip_file:
        with zip_file.open("xlsx_example.xlsx") as xlsx_file:
            workbook = load_workbook(xlsx_file)
            workbook_content = workbook.active

            assert workbook_content.max_row == 51
            assert workbook_content.max_column == 8
